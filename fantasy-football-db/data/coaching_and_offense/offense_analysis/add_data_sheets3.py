import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ARIAL="Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=ARIAL, size=10)

wb = openpyxl.load_workbook('Decline_Adjusted_and_Alumni_Analysis.xlsx')

def add_df_sheet(name, df, widths=None, pct_cols=None, dec_cols=None):
    ws = wb.create_sheet(name)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col)
    for c in ws[1]:
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            colname = df.columns[j-1]
            if pct_cols and colname in pct_cols:
                c.number_format = '0.0%'
            if dec_cols and colname in dec_cols:
                c.number_format = dec_cols[colname]
    ws.freeze_panes = "A2"
    if widths:
        for j,w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    else:
        for j in range(1, len(df.columns)+1):
            ws.column_dimensions[get_column_letter(j)].width = 15
    return ws

# 1. QB Experience Curve
curve = pd.read_csv('qb_experience_curve.csv')
curve.columns = ['Years of Experience','N (League-Wide)','Avg Offensive Percentile']
add_df_sheet('QB Experience Curve', curve, widths=[16,16,20], pct_cols={'Avg Offensive Percentile'})

# 2. Wilson & Darnold Detail
panel = pd.read_csv('qb_panel_with_residual.csv')
sub = panel[panel.primary_qb_name.isin(['Russell Wilson','Sam Darnold'])].sort_values(['primary_qb_name','season'])
sub = sub[['primary_qb_name','season','team','experience_year','epa_per_play_pctile','expected_pctile','residual_pctile']]
sub.columns = ['QB','Season','Team','Experience Yr','Actual Off. Pctile','Expected Off. Pctile','Residual']
add_df_sheet('Wilson & Darnold Detail', sub, widths=[15,8,7,13,16,18,10],
             pct_cols={'Actual Off. Pctile','Expected Off. Pctile','Residual'})

# 3. Residual-Adjusted Elevation Leaderboard
lb = pd.read_csv('coach_qb_elevation_leaderboard_residual.csv')
lb = lb.rename(columns={'coach':'Coach','n_qbs':'N Distinct QBs','avg_raw_delta':'Avg Raw Delta',
    'avg_resid_delta':'Avg Residual (Decline-Adj.) Delta','t_resid':'T-Stat','p_resid':'P-Value'})
add_df_sheet('Residual-Adj Elevation LB', lb, widths=[18,12,13,26,9,10],
             pct_cols={'Avg Raw Delta','Avg Residual (Decline-Adj.) Delta'}, dec_cols={'T-Stat':'0.00','P-Value':'0.0000'})

# 4. QB Alumni Before-After (Filtered)
alumni = pd.read_csv('qb_alumni_before_after_filtered.csv')
alumni = alumni.rename(columns={'coach':'Coach','qb_name':'QB','n_before':'N Before','avg_resid_before':'Avg Residual Before',
    'n_after':'N After','avg_resid_after':'Avg Residual After','after_minus_before':'After Minus Before'})
alumni = alumni.sort_values('After Minus Before', ascending=False)
add_df_sheet('QB Alumni Before-After', alumni, widths=[18,18,9,16,8,16,15],
             pct_cols={'Avg Residual Before','Avg Residual After','After Minus Before'})

# 5. Coach Alumni Leaderboard
cal = pd.read_csv('coach_alumni_leaderboard.csv')
cal = cal.rename(columns={'coach':'Coach','n_qbs':'N Qualifying QBs','avg_swing':'Avg After-Minus-Before Swing','t':'T-Stat','p':'P-Value'})
add_df_sheet('Coach Alumni Leaderboard', cal, widths=[18,14,22,9,10], pct_cols={'Avg After-Minus-Before Swing'}, dec_cols={'T-Stat':'0.00','P-Value':'0.0000'})

wb.save('Decline_Adjusted_and_Alumni_Analysis.xlsx')
print(wb.sheetnames)
