import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=ARIAL, size=10)

wb = openpyxl.load_workbook('Coaching_and_Offense_Quality_Analysis.xlsx')

def add_df_sheet(name, df, widths=None, pct_cols=None, dec_cols=None):
    ws = wb.create_sheet(name)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col)
    for c in ws[1]:
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            colname = df.columns[j-1]
            if pct_cols and colname in pct_cols:
                c.number_format = '0.0%'
            if dec_cols and colname in dec_cols:
                c.number_format = dec_cols[colname]
    ws.freeze_panes = "A2"
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    else:
        for j in range(1, len(df.columns)+1):
            ws.column_dimensions[get_column_letter(j)].width = 15
    return ws

# 1. Coach Season Detail
csd = pd.read_csv('coach_season_detail.csv')
csd = csd[['coach_name','season','team','role','ppg','epa_per_play','epa_rank','epa_per_play_pctile','ppg_rank']]
csd.columns = ['Coach','Season','Team','Role','PPG','EPA/Play','EPA Rank','EPA Percentile','PPG Rank']
add_df_sheet('Coach Season Detail', csd, widths=[16,9,7,7,9,10,9,12,9], pct_cols={'EPA Percentile'}, dec_cols={'PPG':'0.00','EPA/Play':'0.000'})

# 2. Coach Before-During-After
bda = pd.read_csv('coach_before_during_after_detail.csv')
bda = bda.rename(columns={
    'coach':'Coach','team':'Team','start':'Start Yr','end':'End Yr',
    'during_epa_rank_avg':'During Avg EPA Rank','during_epa_pctile_avg':'During Avg EPA %ile',
    'during_ppg_rank_avg':'During Avg PPG Rank','before_epa_rank':'Before EPA Rank',
    'before_epa_pctile':'Before EPA %ile','after_epa_rank':'After EPA Rank','after_epa_pctile':'After EPA %ile',
    'qb_same_before_to_start':'Same QB (Before->Start)','qb_same_end_to_after':'Same QB (End->After)',
    'n_seasons':'N Seasons'})
add_df_sheet('Coach Before-During-After', bda, widths=[16,7,8,7,14,13,14,11,11,10,10,16,16,10],
             pct_cols={'During Avg EPA %ile','Before EPA %ile','After EPA %ile'})

# 3. Coach Summary
cs = pd.read_csv('coach_summary_pctile.csv')
cs = cs.rename(columns={'coach':'Coach','total_seasons':'Total Seasons (OC+HC combined)',
    'during_epa_pctile_avg':'During Avg EPA %ile','before_epa_pctile_avg':'Before Avg EPA %ile',
    'n_before_obs':'N Before Obs','after_epa_pctile_avg':'After Avg EPA %ile','n_after_obs':'N After Obs'})
ws = add_df_sheet('Coach Summary', cs, widths=[16,16,15,15,10,14,10],
             pct_cols={'During Avg EPA %ile','Before Avg EPA %ile','After Avg EPA %ile'})
# append stat test notes below the table
r = len(cs)+3
notes = [
 "Statistical tests (see Summary sheet Result 1-4 for full narrative):",
 "  Coach-level one-sample t-test, During %ile vs 0.5 (n=6): t=6.747, p=0.0011",
 "  Segment-level paired t-test, During vs Before (n=18): t=3.752, p=0.0016",
 "  Segment-level paired t-test, During vs Before, SAME-QB-ONLY subset (n=11): t=1.939, p=0.0812",
 "  Segment-level paired t-test, During vs After (n=12): t=0.179, p=0.8612",
]
for note in notes:
    ws.cell(row=r, column=1, value=note).font = BODY_FONT
    r += 1

# 4. Player YoY correlation detail
yoy = pd.read_csv('player_yoy_offense_change.csv')
corr_rows = []
from scipy import stats
for pos, g in yoy.groupby('position_y0'):
    r1,p1 = stats.pearsonr(g['off_pctile_change'], g['player_ppg_change'])
    r2,p2 = stats.pearsonr(g['off_pctile_change'], g['player_tier_change'])
    corr_rows.append({'Position':pos,'N Same-Team Year-Pairs':len(g),
        'Corr(Off %ile Chg, PPG Chg)':r1,'p-value (PPG)':p1,
        'Corr(Off %ile Chg, Tier Chg)':r2,'p-value (Tier)':p2})
corr_df = pd.DataFrame(corr_rows).sort_values('Corr(Off %ile Chg, Tier Chg)', ascending=False)
add_df_sheet('Player YoY Offense-Change Corr', corr_df, widths=[10,20,22,12,22,12],
             dec_cols={'Corr(Off %ile Chg, PPG Chg)':'0.000','p-value (PPG)':'0.0000',
                       'Corr(Off %ile Chg, Tier Chg)':'0.000','p-value (Tier)':'0.0000'})

# Also add the underlying share-below-average table
share_rows = [
    {'Position':'QB','% of Star+ seasons on below-avg offense':0.121,'% on bottom-third offense':0.058},
    {'Position':'RB','% of Star+ seasons on below-avg offense':0.341,'% on bottom-third offense':0.215},
    {'Position':'WR','% of Star+ seasons on below-avg offense':0.290,'% on bottom-third offense':0.136},
    {'Position':'TE','% of Star+ seasons on below-avg offense':0.340,'% on bottom-third offense':0.196},
]
share_df = pd.DataFrame(share_rows)
ws2 = add_df_sheet('Player YoY Offense-Change Corr', share_df, widths=[10,32,20]) if False else None
# append below existing sheet instead
wsx = wb['Player YoY Offense-Change Corr']
startr = len(corr_df) + 3
wsx.cell(row=startr, column=1, value="Share of Star-or-better player-seasons on a below-average / bottom-third team offense:").font = BODY_FONT
startr += 1
for j, col in enumerate(share_df.columns, start=1):
    c = wsx.cell(row=startr, column=j, value=col); c.font = HEADER_FONT; c.fill = HEADER_FILL
startr += 1
for _, row in share_df.iterrows():
    for j, val in enumerate(row, start=1):
        c = wsx.cell(row=startr, column=j, value=val)
        c.font = BODY_FONT
        if j > 1:
            c.number_format = '0.0%'
    startr += 1

# 5. Player Examples
df = pd.read_csv('player_season_with_offense_rank.csv')
ex = df[(df.tier_score>=4) & (df.off_rank>=28)].sort_values(['position','off_rank'])
ex = ex[['season','display_name','position','team','off_rank','tier_base','ppg']]
ex.columns = ['Season','Player','Position','Team','Off. Rank (of ~32)','Tier','PPG']
add_df_sheet('Player Examples', ex, widths=[9,20,9,7,17,12,9], dec_cols={'PPG':'0.00'})

# 6. Vegas vs Offense by Era
vv = pd.read_csv('vegas_vs_offense.csv')
vv = vv[['season','team','win_total_line','actual_wins','epa_per_play','epa_rank','epa_per_play_pctile','ppg']]
vv.columns = ['Season','Team','Preseason Win Total Line','Actual Wins','EPA/Play','EPA Rank','EPA Percentile','PPG']
add_df_sheet('Vegas vs Offense by Era', vv, widths=[9,7,20,11,10,9,13,9],
             pct_cols={'EPA Percentile'}, dec_cols={'EPA/Play':'0.000','PPG':'0.00'})

wb.save('Coaching_and_Offense_Quality_Analysis.xlsx')
print(wb.sheetnames)
