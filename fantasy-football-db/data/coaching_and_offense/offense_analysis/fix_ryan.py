import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ARIAL="Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=ARIAL, size=10)
TITLE_FONT = Font(name=ARIAL, bold=True, size=14)

wb = openpyxl.load_workbook('Decline_Adjusted_and_Alumni_Analysis.xlsx')

# Rename and extend the detail sheet to include Matt Ryan
old_ws = wb['Wilson & Darnold Detail']
wb.remove(old_ws)

panel = pd.read_csv('qb_panel_with_residual.csv')
sub = panel[panel.primary_qb_name.isin(['Russell Wilson','Sam Darnold','Matt Ryan'])].sort_values(['primary_qb_name','season'])
sub = sub[['primary_qb_name','season','team','experience_year','epa_per_play_pctile','expected_pctile','residual_pctile']]
sub.columns = ['QB','Season','Team','Experience Yr','Actual Off. Pctile','Expected Off. Pctile','Residual']

ws = wb.create_sheet('Wilson, Darnold & Ryan Detail', 2)
for j, col in enumerate(sub.columns, start=1):
    ws.cell(row=1, column=j, value=col)
for c in ws[1]:
    c.font = HEADER_FONT; c.fill = HEADER_FILL
for i, row in enumerate(sub.itertuples(index=False), start=2):
    for j, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=j, value=val)
        c.font = BODY_FONT
        if j in (5,6,7):
            c.number_format = '0.0%'
ws.freeze_panes = "A2"
for j,w in enumerate([15,8,7,13,16,18,10], start=1):
    ws.column_dimensions[get_column_letter(j)].width = w

# Add a coach-transition annotation column note as a second small table below
r = len(sub)+3
notes = [
 "Matt Ryan coach timeline at Atlanta (for reference):",
 "  2008-2014: pre-Shanahan (various OCs)",
 "  2015-2016: Kyle Shanahan, OC",
 "  2017-2018: Steve Sarkisian, OC (Shanahan gone, Ryan still ATL)",
 "  2019-2020: Dirk Koetter, OC (still ATL)",
 "  2021: Arthur Smith/Dave Ragone, new HC/OC regime (still ATL, Ryan's last ATL season)",
 "  2022: traded to Indianapolis (Frank Reich, HC) -- his first season with a different team",
]
for note in notes:
    ws.cell(row=r, column=1, value=note).font = BODY_FONT
    r += 1

# Fix the Summary sheet's Matt Ryan claim
sws = wb['Summary']
for row in sws.iter_rows():
    for cell in row:
        if cell.value and 'Matt Ryan' in str(cell.value) and 'Shanahan' in str(cell.value):
            print('Found at row', cell.row, ':', cell.value)

wb.save('Decline_Adjusted_and_Alumni_Analysis.xlsx')
print(wb.sheetnames)
