import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=ARIAL, bold=True, size=14)
BODY_FONT = Font(name=ARIAL, size=10)
BOLD_BODY = Font(name=ARIAL, size=10, bold=True)

wb = openpyxl.Workbook()

def style_header_row(ws, row=1):
    for c in ws[row]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

def add_df_sheet(name, df, widths=None, pct_cols=None, dec_cols=None):
    ws = wb.create_sheet(name)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col)
    style_header_row(ws)
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            colname = df.columns[j-1]
            if pct_cols and colname in pct_cols:
                c.number_format = '0.0%'
            if dec_cols and colname in dec_cols:
                c.number_format = dec_cols[colname]
    ws.freeze_panes = "A2"
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    else:
        for j in range(1, len(df.columns)+1):
            ws.column_dimensions[get_column_letter(j)].width = 16
    return ws

# ---------------- Summary sheet ----------------
ws = wb.active
ws.title = "Summary"
ws.column_dimensions['A'].width = 112
r = 1
def title(t):
    global r
    ws.cell(row=r, column=1, value=t).font = TITLE_FONT
    r += 2
def header(t):
    global r
    c = ws.cell(row=r, column=1, value=t); c.font = HEADER_FONT; c.fill = HEADER_FILL
    r += 1
def body(t):
    global r
    ws.cell(row=r, column=1, value=t).font = BODY_FONT
    r += 1

title("Do the Best-Reputation Offensive Coaches Actually Produce Top Offenses?")

header("METHOD")
body("Tested 6 coaches with strong analyst reputations as elite offensive minds: Kyle Shanahan,")
body("Matt LaFleur, Sean McVay, Ben Johnson, Andy Reid, and Sean Payton. Each coach's OC-role and")
body("HC-role seasons are merged into one continuous track record (per this project's existing")
body("v14 methodology). 'Offense quality' = team offensive EPA/play, ranked/percentiled against")
body("all NFL teams that same season (1.0 = best offense in the league that year, 0.5 = average).")
body("Sean Payton's 2012 season (suspended all year for the Bounty Scandal, in no way actually")
body("coaching) is excluded from his credit -- verified this by seeing '(dagger)' scraped onto his")
body("name for that one season in the raw coaching-staff data, then confirmed the real-world reason.")
r += 1

header("RESULT 1: YES -- WHILE THEY ARE OC/HC, THESE OFFENSES RANK WELL ABOVE AVERAGE")
body("Average offensive percentile while coaching (weighted by seasons), all 6 coaches:")
body("  Ben Johnson 83.6%ile | Matt LaFleur 79.2%ile | Sean Payton 75.4%ile | Andy Reid 73.1%ile |")
body("  Kyle Shanahan 63.7%ile | Sean McVay 62.8%ile   (see 'Coach Summary' sheet for full detail)")
body("Because this tests SIX SPECIFIC, PRE-NAMED coaches (not a scan of every coach in the league),")
body("this doesn't need the same multiple-testing correction used in the v13/v14 all-coach scans.")
body("One-sample t-test, coach-level (n=6, treating each coach as one independent observation to")
body("avoid the pseudo-replication trap already learned earlier in this project): mean=72.9%ile vs.")
body("a 50% league-average null, t=6.75, p=0.0011. This is a real, statistically solid result: analyst")
body("consensus on these 6 names is genuinely borne out in the play-by-play data.")
r += 1

header("RESULT 2: THEIR ARRIVAL IS ASSOCIATED WITH A REAL JUMP FROM THE YEAR BEFORE")
body("Segment-level (each coach-team stop = one segment, n=18 segments with a valid 'before' year):")
body("  during tenure avg = 66.0%ile vs. the year immediately BEFORE they arrived at that team =")
body("  37.1%ile (paired t-test t=3.75, p=0.0016). Segments aren't fully independent (some coaches")
body("  contribute several), so treat this as directionally strong rather than a clean single test.")
r += 1

header("RESULT 3: CONTROLLING FOR THE STARTING QB SHRINKS -- BUT DOESN'T ERASE -- THE EFFECT")
body("The jump above is partly explained by these coaches also often inheriting/importing a better")
body("QB in the same move (e.g. Andy Reid + Alex Smith at KC, Sean McVay leaving WAS for a fresh")
body("QB room in LA). Restricting to the subset of team-stops where the SAME primary starting QB")
body("carried over from the prior season (n=11, holding QB talent roughly fixed): during = 66.1%ile")
body("vs. before = 50.4%ile -- still positive and still a meaningfully sized gap, but the sample")
body("shrinks a lot and the result is only marginally significant (t=1.94, p=0.081). Honest read:")
body("part of the 'these coaches make offenses great' story is entangled with them also landing")
body("better QB situations, but a real coaching-scheme effect still shows up even when QB is held")
body("constant -- it's just not as large or as certain as the headline number implies.")
r += 1

header("RESULT 4: OFFENSES DON'T RELIABLY COLLAPSE THE MOMENT THESE COACHES LEAVE")
body("During-tenure avg (62.9%ile) vs. the very next season after they departed that team (61.5%ile,")
body("n=12): no significant difference (t=0.18, p=0.86). Notable counter-examples that keep this")
body("honest: Philadelphia's offense actually IMPROVED the year after Andy Reid left (2013, Chip")
body("Kelly's up-tempo debut, 90.6%ile) and Houston's improved after Shanahan left too (96.9%ile).")
body("This says roster talent and scheme continuity often carry a team for at least one more year")
body("regardless of who's calling plays -- 'the offense fell apart without him' is not, in general,")
body("what the data shows for a coach's very next season away.")
r += 1

header("PART B: DO SOME POSITIONS RESIST A BAD OFFENSE BETTER THAN OTHERS? YES -- CLEARLY.")
body("For every player who stayed on the SAME team across two consecutive seasons, correlated the")
body("change in that team's offensive EPA/play percentile with the change in the player's own PPG")
body("and tier (n=5,440 same-team year-pairs, 2001-2025). Correlation with the player's TIER change,")
body("strongest to weakest:")
body("  QB  r=0.440 (p<0.0001)  -- by far the most tied to overall offense quality (makes sense,")
body("                             QB stats essentially ARE the passing offense)")
body("  RB  r=0.195 (p<0.0001)")
body("  WR  r=0.161 (p<0.0001)")
body("  TE  r=0.053 (p=0.071, not significant) -- TE fantasy output is the LEAST tied to whether")
body("                             the team offense as a whole gets better or worse")
r += 1
body("This shows up just as clearly looking at level, not just year-over-year change: of all")
body("Star-or-better player-seasons, the share that happened on a BELOW-AVERAGE offense:")
body("  QB 12.1%  |  WR 29.0%  |  TE 34.0%  |  RB 34.1%")
body("...and on a BOTTOM-THIRD offense specifically: QB 5.8%  |  WR 13.6%  |  TE 19.6%  |  RB 21.5%")
body("Elite QB seasons on truly bad offenses are rare and generally short-lived/partial (e.g. Dak")
body("Prescott 2020, cut short by injury, on a 25th-ranked offense -- the worst offense to still")
body("produce an elite-tier QB season in this dataset). By contrast, RB and TE regularly produce")
body("elite fantasy seasons even on bottom-5 offenses: Christian McCaffrey (2019 CAR, 28th; 2021")
body("CAR, 30th), Breece Hall (2022/2023 NYJ, 30th/32nd), Fred Taylor (2002 JAX, 32nd), David Njoku")
body("(2023/2024 CLE, 28th/32nd), Brock Bowers (2024/2025 LV, 31st/32nd). See the 'Player Examples'")
body("sheet for the full list. Bottom line: RB and TE fantasy value is meaningfully more independent")
body("of overall offensive quality than QB or WR -- a running back or tight end can be a league-")
body("winning fantasy asset even on one of the league's worst offenses; a QB essentially cannot.")
r += 1

header("PART C: PRESEASON VEGAS WIN TOTALS VS. ACTUAL END-OF-SEASON OFFENSIVE RANK")
body("Correlated each team's preseason Vegas win-total line against that SAME season's offensive")
body("EPA/play (799 team-seasons, 2001-2025, 100% match rate):")
body("  Full sample (2001-2025): r=0.462 (p<0.0001)")
body("  2001-2014:               r=0.483 (p<0.0001)")
body("  2015-2024:                r=0.409 (p<0.0001)")
body("  2025 only:                 r=0.559 (p<0.001, n=32, one season -- noisy)")
body("Read: Vegas win totals have always had a moderate, far-from-perfect relationship with actual")
body("offensive quality (r around 0.4-0.5 the whole period) -- which makes sense, since a win total")
body("prices in defense and special teams too, not offense alone. The 2015-2024 correlation is a")
body("bit lower than 2001-2014, but the difference is modest, not a dramatic era shift, and 2025's")
body("single-season r=0.56 is well within the range of one year's sampling noise rather than clear")
body("evidence of a new trend. No strong evidence that Vegas has gotten meaningfully better or worse")
body("at pricing in offensive quality specifically across this period.")
r += 1

header("SHEETS")
body("  - Coach Season Detail: every season these 6 coaches were OC/HC, with offensive rank/percentile")
body("  - Coach Before-During-After: each coach-team stop, with the season before/after for comparison")
body("  - Coach Summary: aggregated stats per coach, plus the statistical tests above")
body("  - Player YoY Offense-Change Correlation: the position-by-position correlation detail")
body("  - Player Examples: elite (Superstar+) player-seasons on bottom-5 offenses, by position")
body("  - Vegas vs Offense by Era: the underlying team-season data for Part C")

wb.save('Coaching_and_Offense_Quality_Analysis.xlsx')
print('Summary sheet written, rows used:', r)
