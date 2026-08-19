import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=ARIAL, size=10)

wb = openpyxl.load_workbook('QB_Controlled_Coach_Analysis.xlsx')

def add_df_sheet(name, df, widths=None, pct_cols=None, dec_cols=None):
    ws = wb.create_sheet(name)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col)
    for c in ws[1]:
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            colname = df.columns[j-1]
            if pct_cols and colname in pct_cols:
                c.number_format = '0.0%'
            if dec_cols and colname in dec_cols:
                c.number_format = dec_cols[colname]
    ws.freeze_panes = "A2"
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    else:
        for j in range(1, len(df.columns)+1):
            ws.column_dimensions[get_column_letter(j)].width = 15
    return ws

named = ['Kyle Shanahan','Matt LaFleur','Sean McVay','Ben Johnson','Andy Reid','Sean Payton']

# 1. Named-6 QB Variety
detail = pd.read_csv('coach_qb_variety_detail.csv')
sub = detail[detail.coach.isin(named)].sort_values(['coach','start']).copy()
sub = sub[['coach','team','start','end','n_seasons','n_distinct_qbs','avg_off_pctile','qb_list']]
sub.columns = ['Coach','Team','Start','End','Seasons','Distinct QBs','Avg Off. Percentile','QB List (Seasons Each)']
add_df_sheet('Named-6 QB Variety', sub, widths=[15,7,7,7,9,11,15,55], pct_cols={'Avg Off. Percentile'})

# 2. Named-6 QB Elevation Detail
pairs = pd.read_csv('qb_coach_paired_deltas_v2.csv')
sub2 = pairs[pairs.coach.isin(named)].sort_values(['coach','delta'], ascending=[True,False]).copy()
sub2 = sub2[['coach','qb_name','n_seasons_with','pctile_with','n_seasons_elsewhere','pctile_elsewhere','delta']]
sub2.columns = ['Coach','QB','Seasons With Coach','Avg Pctile With','Seasons Elsewhere (Career)','Avg Pctile Elsewhere','Delta']
add_df_sheet('Named-6 QB Elevation Detail', sub2, widths=[15,18,15,13,20,16,9],
             pct_cols={'Avg Pctile With','Avg Pctile Elsewhere','Delta'})

# 3. All-Coach QB Elevation Leaderboard
lb = pd.read_csv('coach_qb_elevation_leaderboard_v2.csv')
lb = lb.rename(columns={'coach':'Coach','n_qbs':'N Distinct QBs','avg_delta':'Avg Delta',
    'avg_pctile_with':'Avg Pctile With','avg_pctile_elsewhere':'Avg Pctile Elsewhere',
    't_stat':'T-Stat','p_value':'P-Value'})
add_df_sheet('All-Coach QB Elevation LB', lb, widths=[18,12,10,14,16,9,10],
             pct_cols={'Avg Delta','Avg Pctile With','Avg Pctile Elsewhere'}, dec_cols={'T-Stat':'0.00','P-Value':'0.0000'})

# 4. Career Avg Offensive Percentile (All Coaches)
career = pd.read_csv('coach_career_avg_pctile.csv')
career = career.rename(columns={'coach':'Coach','total_seasons':'Total Seasons (OC+HC, All Teams)','avg_pctile':'Career Avg Off. Percentile'})
add_df_sheet('Career Avg Off Pctile (All)', career, widths=[18,22,20], pct_cols={'Career Avg Off. Percentile'})

# 5. Combined Leaderboard
comb = pd.read_csv('coach_combined_leaderboard.csv')
comb = comb.rename(columns={'coach':'Coach','total_seasons':'Total Seasons','avg_pctile':'Career Avg Off. Percentile',
    'n_qbs':'N Distinct QBs','avg_delta':'Avg QB Elevation Delta','avg_pctile_with':'Avg Pctile With',
    'avg_pctile_elsewhere':'Avg Pctile Elsewhere','t_stat':'T-Stat','p_value':'P-Value'})
add_df_sheet('Combined Leaderboard', comb, widths=[18,11,20,12,17,13,15,8,9],
             pct_cols={'Career Avg Off. Percentile','Avg QB Elevation Delta','Avg Pctile With','Avg Pctile Elsewhere'},
             dec_cols={'T-Stat':'0.00','P-Value':'0.0000'})

wb.save('QB_Controlled_Coach_Analysis.xlsx')
print(wb.sheetnames)
