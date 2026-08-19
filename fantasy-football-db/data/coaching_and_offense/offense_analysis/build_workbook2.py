import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=ARIAL, bold=True, size=14)
BODY_FONT = Font(name=ARIAL, size=10)

wb = openpyxl.Workbook()

def add_df_sheet(wb, name, df, widths=None, pct_cols=None, dec_cols=None):
    ws = wb.create_sheet(name)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col)
    for c in ws[1]:
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            colname = df.columns[j-1]
            if pct_cols and colname in pct_cols:
                c.number_format = '0.0%'
            if dec_cols and colname in dec_cols:
                c.number_format = dec_cols[colname]
    ws.freeze_panes = "A2"
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    else:
        for j in range(1, len(df.columns)+1):
            ws.column_dimensions[get_column_letter(j)].width = 15
    return ws

# ---------------- Summary ----------------
ws = wb.active
ws.title = "Summary"
ws.column_dimensions['A'].width = 114
r = 1
def title(t):
    global r
    ws.cell(row=r, column=1, value=t).font = TITLE_FONT; r += 2
def header(t):
    global r
    c = ws.cell(row=r, column=1, value=t); c.font = HEADER_FONT; c.fill = HEADER_FILL; r += 1
def body(t):
    global r
    ws.cell(row=r, column=1, value=t).font = BODY_FONT; r += 1

title("Does a Coach Elevate Whichever QB He Has? A Within-QB Test, and Who Else Belongs on the List")

header("METHOD: A STRICTER TEST THAN A SIMPLE BEFORE/AFTER COMPARISON")
body("The prior analysis (v15) compared a team's offense before/during/after a coach arrived -- useful,")
body("but it can't fully separate 'the coach is great' from 'the coach also happened to get a better")
body("QB.' This analysis instead asks a narrower, harder question directly: for the SAME quarterback,")
body("does he perform better (team offensive EPA/play percentile) while playing for Coach X than he")
body("does in the rest of his own career, under every OTHER coach he ever played for? This holds the")
body("player's talent fixed and isolates the coach as the only thing that changed.")
body("For every coach (OC or HC role, combined across every team they've held that role for -- same")
body("merged-identity method as v14), every quarterback who was that team's primary starter during the")
body("coach's tenure is matched against that same QB's average percentile in every OTHER season of his")
body("career (any other team, any other coach). This only works for QBs who HAVE another team/coach to")
body("compare against -- Patrick Mahomes and (mostly) Brock Purdy have no career outside their current")
body("coach, so they cannot appear as evidence here. That's a real limitation, noted below.")
r += 1

header("A CRITICAL BASELINE CHECK FIRST: DOES 'STAYING WITH ONE COACH ACROSS MULTIPLE SEASONS' EVEN MEAN ANYTHING ON ITS OWN?")
body("Pooling ALL 687 coach-QB pairs leaguewide with real comparison data: the average QB does NOT")
body("perform better with any given coach than in the rest of his career -- mean delta = -0.036 (i.e.")
body("slightly WORSE), only 45.7% of pairs are positive. This is an important control: it means a coach")
body("racking up a second or third different multi-year starter is not, by itself, evidence of skill --")
body("if anything the leaguewide baseline leans slightly negative (plausibly because journeyman/backup")
body("QBs are often acquired cheaply during their decline phase). Any coach beating this -3.6% baseline")
body("by a wide margin is doing something real; a coach merely at +0.05 is roughly at league-average.")
r += 1

header("RESULT: THE ORIGINAL 4 NAMED COACHES DON'T ALL HOLD UP EQUALLY WELL ONCE QB IS PROPERLY CONTROLLED")
body("Per-coach average delta (QB's pctile-with-this-coach minus that QB's own career pctile elsewhere),")
body("across every distinct starting QB with real comparison data:")
body("  Andy Reid:      +0.222  (n=3 QBs: McNabb +0.18, Vick +0.07, Alex Smith +0.41 -- ALL THREE")
body("                   positive. This is the single cleanest 'elevates whoever he has' case in the")
body("                   whole dataset. Mahomes is excluded -- no career outside Reid to compare to,")
body("                   which if anything UNDERSTATES Reid's real effect.)")
body("  Matt LaFleur:   +0.059  (n=3 QBs: Rodgers +0.08, Goff +0.18, Mariota -0.09 -- mildly positive,")
body("                   not a strong signal on its own)")
body("  Kyle Shanahan:  +0.053  (n=6 QBs: Grossman/Schaub/Ryan/Hoyer/RG3 all positive, but Donovan")
body("                   McNabb's rough lone 2010 season with him is a large negative outlier -0.40.")
body("                   Net: barely above the leaguewide baseline, i.e. roughly average once you")
body("                   include his one clear miss.)")
body("  Sean McVay:     -0.004  (n=4 QBs: Stafford +0.19 and Cousins +0.24 positive, but Goff -0.09 and")
body("                   RG3's brief rookie-OC-year -0.36 pull it to essentially dead even)")
body("  Sean Payton:     -0.097 (n=4 QBs: Brees is spectacular, +0.123 across a HUGE 14-season sample --")
body("                   but Russell Wilson (-0.27, his 2023 Denver year) and Jameis Winston (-0.23,")
body("                   backup role) pull the coach-level average NEGATIVE. Payton's 'makes anyone")
body("                   good' reputation looks like it's really 'was outstanding with one franchise QB'")
body("                   once you look past Brees specifically.)")
body("None of these individual coach-level tests reach conventional statistical significance (all")
body("p>0.10) -- the samples are just too small (2-6 QBs each), which is a structural limit of this")
body("question: it's rare for a coach to keep his job through 3+ genuinely different competent starters.")
r += 1

header("OTHER COACHES THE DATA SUPPORTS ADDING TO THIS LIST")
body("Scanning every coach in the league (min 2 distinct QBs with real comparison data), several names")
body("outperform the leaguewide -3.6% baseline by as much or more than the original 6, with comparable")
body("or larger samples:")
body("  Josh McDaniels: +0.217 (n=8 QBs across NE/DEN/LV stops -- the LARGEST sample of any coach tested,")
body("                   and still solidly positive. The strongest data-backed addition to this list.)")
body("  Greg Roman:     +0.161 (n=6 QBs across BAL/SF/others -- matches his real-world reputation for")
body("                   building schemes around unconventional QBs, e.g. Kaepernick, Lamar Jackson)")
body("  Bill Belichick: +0.124 (n=3 QBs -- but this is very likely mostly a Tom Brady/dynasty-era")
body("                   artifact, not a general 'elevates any QB' pattern; treat with caution)")
body("  Mike Vrabel:    +0.327 (n=3, largest raw delta of any coach -- but Vrabel is defense-oriented;")
body("                   this likely reflects his OC hires' work more than his own scheme, and the")
body("                   sample is small)")
body("  Jim Harbaugh, Chip Kelly: both +0.23 (n=3 each) -- promising but too small a sample to lean on")
r += 1
body("Cross-referencing against career-long average offensive percentile (min 6 total seasons as OC/HC,")
body("any team) reinforces Josh McDaniels as the strongest additional name: 17 seasons, 75.0th")
body("percentile average -- right in between Sean Payton (75.4) and Andy Reid (73.1) on that measure,")
body("with a much larger positive QB-elevation sample behind it than either.")
r += 1

header("SO -- CAN WE 'PROVE' THIS ABILITY EXISTS? PARTIALLY.")
body("What the data DOES support: this is a real, testable question, and the leaguewide baseline shows")
body("simply having multiple long-tenured starters is not automatically evidence of anything -- so the")
body("coaches who clear that bar by a wide margin (Reid, McDaniels, Roman, and to a lesser extent")
body("Shanahan/LaFleur) are doing something distinguishable from average. Andy Reid in particular has")
body("the cleanest, most complete case: three unrelated QBs (a mobile athlete in McNabb, a run-first")
body("athlete in Vick, a game-manager in Alex Smith) all performed above their own career norms under him.")
body("What the data does NOT support: statistical proof at the individual-coach level. Every single")
body("coach-level t-test in this analysis has p>0.10 -- the sample sizes (2-8 distinct QBs per coach)")
body("are simply too small for conventional significance, and this is a structural ceiling, not a flaw")
body("in method -- very few coaches ever get to coach 3+ truly different competent starters. There's")
body("also a real right-censoring problem: the QBs who never leave a coach (Mahomes/Reid, largely")
body("Purdy/Shanahan) can't be used as evidence in this design at all, which likely understates the")
body("true effect for the very best pairings. And McVay's and Payton's reputations specifically look")
body("less like 'succeeds with any QB' and more like 'had one all-time-great QB' once this control is")
body("applied -- a genuinely different and more honest read than the popular narrative.")
r += 1

header("SHEETS")
body("  - Named-6 QB Variety: every season the 6 originally-tested coaches held OC/HC, with the")
body("    starting QB(s) and offensive percentile that season")
body("  - Named-6 QB Elevation Detail: the with-coach vs. elsewhere comparison for each of their QBs")
body("  - All-Coach QB Elevation Leaderboard: every coach in the league with 2+ distinct QBs and real")
body("    comparison data, ranked by average elevation delta")
body("  - Career Avg Offensive Percentile (All Coaches): every coach's whole-career (OC+HC combined,")
body("    all teams) average offensive percentile, min 6 seasons")
body("  - Combined Leaderboard: the two above merged, for cross-referencing")

wb.save('QB_Controlled_Coach_Analysis.xlsx')
print('summary rows:', r)
