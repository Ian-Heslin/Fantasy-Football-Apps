import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ARIAL="Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=ARIAL, size=10)

wb = openpyxl.load_workbook('Top_Offense_Consistency_and_Superstars.xlsx')

def add_df_sheet(name, df, widths=None, pct_cols=None, dec_cols=None):
    ws = wb.create_sheet(name)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col)
    for c in ws[1]:
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            colname = df.columns[j-1]
            if pct_cols and colname in pct_cols:
                c.number_format = '0.0%'
            if dec_cols and colname in dec_cols:
                c.number_format = dec_cols[colname]
    ws.freeze_panes = "A2"
    if widths:
        for j,w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    else:
        for j in range(1, len(df.columns)+1):
            ws.column_dimensions[get_column_letter(j)].width = 15
    return ws

# 1. Top Offense Consistency
cons = pd.read_csv('coach_top_offense_consistency.csv')
cons = cons.rename(columns={'coach_name':'Coach','total_seasons':'Total OC/HC Seasons',
    'top5_seasons':'Top-5 Seasons','top10_seasons':'Top-10 Seasons','avg_rank':'Avg Offensive Rank',
    'top5_rate':'Top-5 Rate','top10_rate':'Top-10 Rate'})
add_df_sheet('Top Offense Consistency', cons, widths=[18,14,12,13,13,10,11],
             pct_cols={'Top-5 Rate','Top-10 Rate'}, dec_cols={'Avg Offensive Rank':'0.00'})

# 2. Superstar Rate (Combined)
ss = pd.read_csv('coach_superstar_rate_combined.csv')
ss = ss[ss.n_distinct_players>=15].sort_values('superstar_rate', ascending=False)
ss = ss.rename(columns={'coach_name':'Coach','n_distinct_players':'N Distinct Skill Players',
    'n_superstar_players':'N Reached Superstar+','superstar_rate':'Superstar+ Rate'})
add_df_sheet('Superstar Rate (Combined)', ss, widths=[18,18,16,13], pct_cols={'Superstar+ Rate'})

# 3. Merged
m = pd.read_csv('coach_consistency_and_superstar.csv')
m = m[m.n_distinct_players>=15].sort_values('top10_rate', ascending=False)
m = m.rename(columns={'coach_name':'Coach','total_seasons':'OC/HC Seasons','top5_rate':'Top-5 Rate',
    'top10_rate':'Top-10 Rate','avg_rank':'Avg Off. Rank','n_distinct_players':'N Distinct Players',
    'n_superstar_players':'N Superstar+','superstar_rate':'Superstar+ Rate'})
m = m[['Coach','OC/HC Seasons','Top-5 Rate','Top-10 Rate','Avg Off. Rank','N Distinct Players','N Superstar+','Superstar+ Rate']]
ws = add_df_sheet('Consistency + Superstar Merged', m, widths=[18,12,10,11,11,13,10,13],
             pct_cols={'Top-5 Rate','Top-10 Rate','Superstar+ Rate'}, dec_cols={'Avg Off. Rank':'0.00'})

startr = len(m)+3
notes = [
 "Statistical tests (97 coaches, >=6 OC/HC seasons and >=15 distinct players):",
 "  Pearson corr(Top-10 Rate, Superstar+ Rate): r=0.588, p<0.0001",
 "  Spearman corr(Top-10 Rate, Superstar+ Rate): r=0.560, p<0.0001",
 "  Fisher's exact, elite (Top-10 Rate>=60%, n=14 coaches, 1252 players) vs rest (n=83, 7532 players):",
 "    Elite superstar rate 7.6% vs rest 5.2%, odds ratio=1.51, p=0.0008",
 "  Robustness check across thresholds (all consistently OR 1.4-1.6, p<0.02):",
 "    Top-10 Rate>=40%: OR=1.40, p=0.00043  |  >=50%: OR=1.56, p=0.00002  |  >=70%: OR=1.62, p=0.00214",
 "    Top-5 Rate>=30%: OR=1.56, p=0.00036  |  >=40%: OR=1.49, p=0.00482  |  >=50%: OR=1.57, p=0.01294",
]
for note in notes:
    ws.cell(row=startr, column=1, value=note).font = BODY_FONT
    startr += 1

wb.save('Top_Offense_Consistency_and_Superstars.xlsx')
print(wb.sheetnames)
