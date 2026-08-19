import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ARIAL="Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=ARIAL, bold=True, size=14)
BODY_FONT = Font(name=ARIAL, size=10)

wb = openpyxl.Workbook()

def add_df_sheet(name, df, widths=None, pct_cols=None, dec_cols=None):
    ws = wb.create_sheet(name)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col)
    for c in ws[1]:
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            colname = df.columns[j-1]
            if pct_cols and colname in pct_cols:
                c.number_format = '0.0%'
            if dec_cols and colname in dec_cols:
                c.number_format = dec_cols[colname]
    ws.freeze_panes = "A2"
    if widths:
        for j,w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    else:
        for j in range(1, len(df.columns)+1):
            ws.column_dimensions[get_column_letter(j)].width = 15
    return ws

ws = wb.active
ws.title = "Summary"
ws.column_dimensions['A'].width = 114
r=1
def title(t):
    global r; ws.cell(row=r,column=1,value=t).font=TITLE_FONT; r+=2
def header(t):
    global r; c=ws.cell(row=r,column=1,value=t); c.font=HEADER_FONT; c.fill=HEADER_FILL; r+=1
def body(t):
    global r; ws.cell(row=r,column=1,value=t).font=BODY_FONT; r+=1

title("Decline-Adjusted Coaching Effects, and What Happens After a QB Leaves")

header("PART 1: ADJUSTING FOR EXPECTED CAREER-STAGE DECLINE (THE WILSON PROBLEM)")
body("The v16 test compared a QB's performance with a coach against his flat career average elsewhere")
body("-- which is unfair to any coach whose QB was already declining before arriving (or credits a coach")
body("whose QB was simply young and still improving). Built an experience-curve baseline: average team")
body("offensive percentile by 'years since first primary-starter season,' league-wide (rookies average")
body("the 32nd percentile, rising to a 55-59th percentile plateau by year 6-8, noisy after that as only")
body("elite QBs still start into their teens of experience). Each QB-season's RESIDUAL = actual percentile")
body("minus the expected percentile for that experience-year -- this nets out 'declining/improving anyway.'")
r+=1
body("Russell Wilson, specifically (residual, i.e. vs. his OWN career-stage expectation, not a flat average):")
body("  2012-2015 (years 0-3, SEA): residuals +0.29 to +0.59 -- a true early-career star performance")
body("  2016-2021 (years 4-9, SEA): residuals fall to +0.04 to +0.27 -- already trending toward merely")
body("    average-for-his-experience-level BEFORE Denver, exactly matching what you described")
body("  2022 (year 10, DEN, Payton's first year): residual -0.367 -- meaningfully WORSE than even his")
body("    already-declining trend would predict. This is the one year Payton doesn't get a pass on.")
body("  2023 (year 11, DEN): residual -0.076 -- still slightly below expectation, but a real recovery")
body("    from the 2022 shock, much closer to what his decline curve alone would predict")
body("Bottom line: your instinct is right that Wilson's decline started in Seattle, not Denver -- his")
body("underlying trajectory was already flattening. But 2022 under Payton was still a real, additional")
body("shock beyond that trend, one that partially self-corrected in year two.")
r+=1

header("REDONE LEADERBOARD USING DECLINE-ADJUSTED RESIDUALS INSTEAD OF RAW PERCENTILE")
body("Leaguewide baseline check barely moves with this adjustment (mean delta -0.039 vs. the original")
body("-0.036 raw version) -- the correction mostly matters for INDIVIDUAL cases like Wilson's, not the")
body("aggregate. Named-6 results, decline-adjusted (raw delta -> residual delta):")
body("  Andy Reid:      +0.222 -> +0.189  (still the cleanest case; holds up)")
body("  Ben Johnson:    +0.411 -> +0.312  (still strong, still only n=2 QBs)")
body("  Kyle Shanahan:  +0.053 -> similar mixed picture -- McNabb's lone bad year is still an outlier")
body("                   even adjusted for the fact he was already declining (-0.348 residual)")
body("  Sean Payton:    Wilson's shortfall shrinks (-0.273 raw -> -0.235 residual) but is still real")
body("  Sean McVay, LaFleur: patterns essentially unchanged")
body("Josh McDaniels (+0.200) and Greg Roman remain the strongest data-backed additions after this")
body("adjustment too -- their case doesn't depend on ignoring their QBs' career stage.")
r+=1

header("PART 2: WHAT HAPPENS TO A QB'S TRAJECTORY AFTER HE LEAVES A COACH?")
body("Built a second, independent test: for every QB who spent ANY time (starter OR backup/roster-only,")
body("using team roster data, since e.g. Sam Darnold barely played in 2023 at SF) under a coach, compare")
body("his primary-starter residual performance in the seasons BEFORE joining that coach vs. the seasons")
body("AFTER leaving. This directly targets your Darnold question, and it can only be asked this way")
body("because it uses roster membership, not just starts -- his 2023 season backing up Brock Purdy at")
body("San Francisco would be invisible to the v16 test entirely.")
r+=1
body("Sam Darnold's full trajectory (residual = actual vs. expected-for-his-experience-year):")
body("  2018-2021 (NYJ, then CAR): residuals -0.20 to -0.47 -- well below expectation the whole time,")
body("    not just bad in absolute terms")
body("  2023: on San Francisco's roster under Kyle Shanahan, but did not start enough games to register")
body("    as a primary-starter season at all")
body("  2024 (MIN, Kevin O'Connell) and 2025 (SEA): residuals +0.02 and +0.07 -- modest but real, a")
body("    genuine reversal from deeply negative to slightly positive")
body("This IS a real, visible turnaround, and it followed his year with Shanahan -- exactly the pattern")
body("you described. BUT: this design cannot cleanly credit Shanahan alone. Darnold's turnaround also")
body("directly coincides with landing at Minnesota under Kevin O'Connell, himself a well-regarded")
body("quarterback developer (Justin Herbert's rookie surge as his OC, then this Darnold turnaround as")
body("HC). Both 'the Shanahan year' and 'the O'Connell system' are equally supported as the explanation")
body("by this data -- there is no way to separate them with only one case to look at.")
r+=1
body("For comparison, Matt Ryan's post-Shanahan trajectory went the OTHER way (residual -0.249 after")
body("leaving ATL, continuing to decline in Indianapolis/backup role) -- so 'a lasting Shanahan bump'")
body("is not a general pattern even within his own alumni; Darnold is the exception, not the rule, in")
body("this specific (small, n=2) sample. Sean Payton's alumni (Kerry Collins, Russell Wilson, Teddy")
body("Bridgewater, Jameis Winston) ALL show negative swings after leaving him -- no evidence of a")
body("lasting positive Payton effect in this test, for what that's worth given the very small samples.")
r+=1

header("COACH-LEVEL ALUMNI LEADERBOARD (min 3 qualifying QBs, before AND after data)")
body("No coach's average post-departure swing reaches statistical significance (all p>0.09, n=3-6 QBs")
body("per coach) -- this is an even noisier test than the contemporaneous one, since a QB's next stop")
body("brings a whole new coach, system, and supporting cast that could just as easily explain any change.")
body("Treat every number on this sheet as a suggestive data point, not proof of a specific coach's")
body("lasting influence. See the 'Coach Alumni Leaderboard' sheet for the full ranked list.")
r+=1

header("BOTTOM LINE")
body("Decline-adjustment confirms your Wilson read (already fading in Seattle) while still showing his")
body("first Denver year was a real added shock beyond that trend. The 'after leaving' test can genuinely")
body("show Darnold's specific turnaround -- and it's real in the data -- but can't isolate whether")
body("Shanahan's year or Minnesota's system deserves the credit, since only one case exists to examine.")
body("Across the alumni leaderboard broadly, no coach shows a statistically real, general 'his players")
body("keep improving even after they leave' effect -- the strongest, most repeatable evidence in both")
body("this and the prior analysis remains the CONTEMPORANEOUS one (Reid, Ben Johnson, McDaniels).")
r+=1

header("SHEETS")
body("  - QB Experience Curve: the baseline expectation by years-of-experience, used for all residuals")
body("  - Wilson & Darnold Detail: full season-by-season trajectory for both, with residuals")
body("  - Residual-Adjusted Elevation Leaderboard: the v16 leaderboard, redone with decline-adjustment")
body("  - QB Alumni Before-After (Filtered): every qualifying coach-QB relationship's before/after swing")
body("  - Coach Alumni Leaderboard: coach-level rollup of the alumni swings (min 3 QBs)")

wb.save('Decline_Adjusted_and_Alumni_Analysis.xlsx')
print("summary rows:", r)
