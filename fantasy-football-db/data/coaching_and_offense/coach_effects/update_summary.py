import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.load_workbook('Coaching_Staff_Player_Development_Analysis.xlsx')
ws = wb['Summary']

# Find existing font/style conventions from an existing populated cell
sample_cell = ws['A1']
title_font = sample_cell.font
header_cell = None
for row in ws.iter_rows():
    for c in row:
        if c.value == 'DATA SOURCE':
            header_cell = c
            break
    if header_cell:
        break
from copy import copy
header_font = copy(header_cell.font)
header_fill = copy(header_cell.fill)
body_font = copy(ws['A4'].font)

# Find first fully empty row after existing content
max_row = ws.max_row
start = max_row + 2

def write_header(r, text):
    c = ws.cell(row=r, column=1, value=text)
    c.font = header_font
    c.fill = header_fill
    return r + 1

def write_body(r, text):
    c = ws.cell(row=r, column=1, value=text)
    c.font = body_font
    return r + 1

r = start
r = write_header(r, "ADDITIONAL ANALYSIS (per user follow-up): DRAFT-PEDIGREE-ADJUSTED COACHING EFFECT")
r = write_body(r, "Question: are some coaches better at getting more out of players who were drafted later, relative to")
r = write_body(r, "how players drafted in that range normally perform? Players were bucketed by draft slot into Day1")
r = write_body(r, "(Round 1), Day2 (Rounds 2-3), and Day3+UDFA (Rounds 4-7 or undrafted), per position. A league baseline")
r = write_body(r, "was built as the average career tier-score for distinct players in each (position, draft tier) cell --")
r = write_body(r, "see 'Draft Tier Baseline' sheet. Each coach's distinct players in a (position, draft tier) cell (min 15")
r = write_body(r, "distinct players) were then compared to the REST of the league's players in that same cell (Welch's")
r = write_body(r, "two-sample t-test, excluding the coach's own players from the comparison group so a coach can't be")
r = write_body(r, "compared partly against himself).")
r = write_body(r, "NOTE ON METHOD REVISION: a first pass (min 5 players, one-sample t-test against the baseline mean)")
r = write_body(r, "produced several p=0.0000 results driven by tiny, zero-variance samples (e.g. 5 players who all")
r = write_body(r, "happened to bust) -- a statistical artifact, not a real finding. This was caught and discarded in favor")
r = write_body(r, "of the larger-sample, two-sample-t-test version described above and shown on the sheets.")
r = write_body(r, "RESULT: NO POSITIVE 'DEVELOPS LATE PICKS BETTER' EFFECT SURVIVES CORRECTION. Across 44 position-coach,")
r = write_body(r, "29 OC, and 28 HC (position x draft-tier) combinations tested, exactly ONE survives Bonferroni correction")
r = write_body(r, "in each table -- and all three are NEGATIVE (underperformance), all at WR Day3+UDFA: Zach Azzanni")
r = write_body(r, "(p=1.0e-07), Greg Roman as OC (p=9.8e-09), and Ron Rivera as HC (p=6.7e-25). In other words, the data")
r = write_body(r, "shows more statistically robust evidence of specific coaches getting LESS out of late-round/UDFA")
r = write_body(r, "receivers than league average, than of any coach getting MORE out of them. No coach clears the bar for")
r = write_body(r, "a genuine 'late-round developer' effect at any position once corrected for the ~100 tests run.")
r += 1

r = write_header(r, "ADDITIONAL ANALYSIS (per user follow-up): COMBINED OC+HC COACHING-IDENTITY EFFECT (Kyle Shanahan example)")
r = write_body(r, "Question, from the user directly: 'I think when we are analyzing coaches we should analyze Kyle Shanahan")
r = write_body(r, "the OC and his impact the same as we analyze Kyle Shanahan the HC.' Per this instruction, each coach's")
r = write_body(r, "OC-role seasons and HC-role seasons were merged into ONE continuous track record (union of all")
r = write_body(r, "team-seasons in either role), rather than scored as separate coaching identities. For every distinct")
r = write_body(r, "skill player (QB/RB/WR/TE) who played under that combined tenure, the player's own overperformance vs.")
r = write_body(r, "draft-tier expectation (actual tier-score minus the position+draft-tier league baseline) was computed,")
r = write_body(r, "then averaged across the coach's distinct players and tested with a one-sample t-test against zero.")
r = write_body(r, "Coaches needed >=15 distinct skill players across their combined tenure to be included (332 qualifying")
r = write_body(r, "coaches).")
r = write_body(r, "RESULT, SCANNING ALL COACHES: 0 of 332 combined coaching identities survive Bonferroni correction. As with")
r = write_body(r, "the earlier superstar/breakout analysis, this reinforces that no coach's whole-career track record is")
r = write_body(r, "statistically distinguishable from luck once compared simultaneously against ~330 other coaches.")
r = write_body(r, "RESULT, AS A SINGLE NAMED HYPOTHESIS (Kyle Shanahan specifically, per the user's question): this is a")
r = write_body(r, "materially different statistical question -- testing ONE pre-specified coach the user named, not scanning")
r = write_body(r, "for the best of many, so no correction for hundreds of comparisons is required. Kyle Shanahan's combined")
r = write_body(r, "9 seasons as OC (HOU 2008-09, WAS 2010-13, CLE 2014, ATL 2015-16) + 9 seasons as HC (SF 2017-2025) shows")
r = write_body(r, "a real, standalone-significant result: +0.262 average overperformance vs. draft-tier expectation across")
r = write_body(r, "120 distinct skill-position players (p=0.0119). The player-level detail (see 'Shanahan Worked Example'")
r = write_body(r, "sheet) directly corroborates the specific examples in the user's question: Robert Griffin III (+1.68,")
r = write_body(r, "consistent with his ROTY season), Matt Ryan (+1.18, consistent with his lone MVP under Shanahan), Brock")
r = write_body(r, "Purdy (mid/late-round QB overperformance), Devonta Freeman (+4.19), George Kittle (+3.83), Christian")
r = write_body(r, "McCaffrey (+2.35), Deebo Samuel (+2.04), Brandon Aiyuk (+0.70, a partial season before injury, as the user")
r = write_body(r, "described), and Arian Foster (+2.69, a UDFA who became a star). This is genuine, specific supporting")
r = write_body(r, "evidence for the user's gut feel about this particular coach -- it does not, however, generalize into")
r = write_body(r, "evidence that coaching effects are detectable league-wide (see the 0/332 scan-all-coaches result above).")
r = write_body(r, "CAVEAT: the 'HC' role was credited broadly, including to defensive-minded head coaches who do not call")
r = write_body(r, "their own offense. Several names near the bottom of the 'Combined OC+HC Rank' sheet (e.g. Rod Marinelli,")
r = write_body(r, "Dick LeBeau, Steve Wilks) are defensive specialists whose 'underperformance' likely reflects that they")
r = write_body(r, "were not the actual offensive play-caller, rather than a real negative effect on player development.")
r += 1

r = write_header(r, "SHEETS (ADDITIONAL)")
r = write_body(r, "  - Draft Tier Baseline: league average career tier-score by position x draft tier (Day1/Day2/Day3+UDFA)")
r = write_body(r, "  - Draft-Tier - Position Coach / OC / HC: coach vs. rest-of-league comparison by position x draft tier")
r = write_body(r, "  - Combined OC+HC Rank: per-coach overperformance vs. draft-tier expectation, OC+HC tenures merged")
r = write_body(r, "  - Shanahan Worked Example: full player-by-player detail for Kyle Shanahan's combined 18-year tenure")

wb.save('Coaching_Staff_Player_Development_Analysis.xlsx')
print('Summary sheet updated, new max_row:', ws.max_row)
