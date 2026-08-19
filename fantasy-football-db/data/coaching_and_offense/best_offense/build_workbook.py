import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=ARIAL, bold=True, size=14)
SUBTITLE_FONT = Font(name=ARIAL, italic=True, size=10, color="444444")
BODY_FONT = Font(name=ARIAL, size=10)
BOLD_BODY = Font(name=ARIAL, size=10, bold=True)

m = pd.read_csv('team_offense_master.csv')
m = m.sort_values(['season','team'])

def best_by(col, name):
    idx = m.groupby('season')[col].idxmax()
    r = m.loc[idx, ['season','team',col]].copy()
    r = r.rename(columns={'team': f'{name}_team', col: f'{name}_value'})
    return r.set_index('season')

ppg = best_by('ppg', 'ppg')
ypg = best_by('yards_per_game', 'ypg')
epa = best_by('epa_per_play', 'epa')
best = ppg.join(ypg).join(epa).reset_index().sort_values('season')

wb = openpyxl.Workbook()

# ---------- Summary sheet ----------
ws = wb.active
ws.title = "Summary"
ws.column_dimensions['A'].width = 110
r = 1
def title(text):
    global r
    c = ws.cell(row=r, column=1, value=text)
    c.font = TITLE_FONT
    r += 2

def header(text):
    global r
    c = ws.cell(row=r, column=1, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    r += 1

def body(text, bold=False):
    global r
    c = ws.cell(row=r, column=1, value=text)
    c.font = BOLD_BODY if bold else BODY_FONT
    r += 1

title("Best NFL Offense by Season, 2000-2025")
header("DATA SOURCE & METHOD")
body("All figures below are computed directly from nflverse play-by-play and schedule data")
body("(play_by_play_{season}.csv, games.csv), regular season only, 2000-2025 (26 seasons).")
body("Three objective, exactly-reproducible metrics are used:")
body("  1) PPG - points per game, from actual final scores (games.csv), regular season.")
body("  2) YPG - yards per game, sum of yards_gained on every offensive scrimmage play")
body("     (pass attempts + rush attempts, excluding penalties/no-plays), divided by games played.")
body("  3) EPA/play - mean Expected Points Added per offensive scrimmage play (pass + rush),")
body("     nflverse's standard advanced efficiency metric. Widely considered a better measure of")
body("     true offensive quality than raw yards, since it accounts for down/distance/situation")
body("     (e.g. garbage-time yards against a prevent defense count for much less).")
r += 1
header("A FOURTH METRIC (DVOA) COULD NOT BE INCLUDED - HERE'S WHY")
body("DVOA (Defense-adjusted Value Over Average) is Football Outsiders' proprietary advanced")
body("metric, generally considered the gold standard alongside EPA/play. Football Outsiders shut")
body("down in 2023; DVOA now lives at FTN Fantasy (ftnfantasy.com), whose full historical archive")
body("(1977-present) is subscriber-only, licensed for personal use, and not bulk-downloadable for")
body("free. Free year-end recap articles exist for individual seasons (e.g. FTN's 'Final 2023 DVOA")
body("Ratings' names the 2023 49ers as the #1 offense by DVOA), but there is no free, complete,")
body("machine-readable historical series to build a reliable year-by-year table from. If you have")
body("or can get FTN Stats/Pro access, I can help pull the historical workbook and merge it in.")
r += 1
header("HEADLINE RESULT: BEST OFFENSE BY YEAR, ALL THREE METRICS")
body("See the 'Best Offense By Year' sheet for the full 2000-2025 table. Notable highlights that")
body("check out against well-known NFL history:")
body("  - 2000 & 2001 St. Louis Rams (\"The Greatest Show on Turf\") - #1 in all three metrics both years")
body("  - 2007 New England Patriots (16-0 season) - #1 in PPG (36.8), YPG, and EPA/play")
body("  - 2013 Denver Broncos (Peyton Manning's record-setting year) - #1 in PPG (37.9) and YPG")
body("  - 2011 & 2020 Green Bay Packers (Aaron Rodgers MVP seasons) - #1 in PPG and EPA/play both years")
body("  - 2018 Kansas City Chiefs (Mahomes' first MVP year) - #1 in PPG, YPG, and EPA/play")
body("  - 2019 & 2024 Baltimore Ravens (Lamar Jackson MVP seasons) - #1 by EPA/play both years")
r += 1
header("A NOTE ON WHY THE THREE METRICS SOMETIMES DISAGREE")
body("PPG, YPG, and EPA/play agree on the same team in about half of all seasons, but diverge in")
body("the rest - this is expected, not an error. PPG rewards scoring efficiently (short fields off")
body("turnovers, red-zone success) more than raw volume. YPG rewards volume and can be inflated by")
body("garbage-time yardage in blowouts or shootouts. EPA/play is the most \"pure\" measure of per-play")
body("quality independent of pace or garbage time, which is why it's generally preferred by")
body("analysts as the best all-around descriptive stat - but it's also the least intuitive to")
body("read directly (values are point-fractions per play, not a familiar counting stat).")
r += 1
header("SHEETS")
body("  - Best Offense By Year: the #1 team by PPG, YPG, and EPA/play for every season 2000-2025")
body("  - All Teams All Years: full underlying data - every team, every season, all three metrics")

# ---------- Best Offense By Year sheet ----------
ws2 = wb.create_sheet("Best Offense By Year")
headers = ["Season", "Best PPG Team", "PPG", "Best YPG Team", "YPG", "Best EPA/Play Team", "EPA/Play"]
for j, h in enumerate(headers, start=1):
    c = ws2.cell(row=1, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
for i, row in best.iterrows():
    rr = i + 2 if False else None
for idx, (_, row) in enumerate(best.iterrows(), start=2):
    ws2.cell(row=idx, column=1, value=int(row['season'])).font = BODY_FONT
    ws2.cell(row=idx, column=2, value=row['ppg_team']).font = BODY_FONT
    c = ws2.cell(row=idx, column=3, value=round(float(row['ppg_value']),2)); c.font = BODY_FONT; c.number_format = '0.00'
    ws2.cell(row=idx, column=4, value=row['ypg_team']).font = BODY_FONT
    c = ws2.cell(row=idx, column=5, value=round(float(row['ypg_value']),1)); c.font = BODY_FONT; c.number_format = '0.0'
    ws2.cell(row=idx, column=6, value=row['epa_team']).font = BODY_FONT
    c = ws2.cell(row=idx, column=7, value=round(float(row['epa_value']),3)); c.font = BODY_FONT; c.number_format = '0.000'
ws2.freeze_panes = "A2"
for j, w in enumerate([8,14,8,14,8,17,10], start=1):
    ws2.column_dimensions[get_column_letter(j)].width = w

# ---------- All Teams All Years sheet ----------
ws3 = wb.create_sheet("All Teams All Years")
cols = ['season','team','games','total_points','ppg','plays','total_yards','yards_per_game','total_epa','epa_per_play']
labels = ['Season','Team','Games','Total Points','PPG','Offensive Plays','Total Yards','Yards Per Game','Total EPA','EPA Per Play']
for j, h in enumerate(labels, start=1):
    c = ws3.cell(row=1, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
mm = m.sort_values(['season','team']).reset_index(drop=True)
for idx, row in mm.iterrows():
    rr = idx + 2
    ws3.cell(row=rr, column=1, value=int(row['season'])).font = BODY_FONT
    ws3.cell(row=rr, column=2, value=row['team']).font = BODY_FONT
    ws3.cell(row=rr, column=3, value=int(row['games'])).font = BODY_FONT
    ws3.cell(row=rr, column=4, value=float(row['total_points'])).font = BODY_FONT
    c = ws3.cell(row=rr, column=5, value=round(float(row['ppg']),2)); c.font = BODY_FONT; c.number_format='0.00'
    ws3.cell(row=rr, column=6, value=int(row['plays'])).font = BODY_FONT
    ws3.cell(row=rr, column=7, value=float(row['total_yards'])).font = BODY_FONT
    c = ws3.cell(row=rr, column=8, value=round(float(row['yards_per_game']),1)); c.font = BODY_FONT; c.number_format='0.0'
    c = ws3.cell(row=rr, column=9, value=round(float(row['total_epa']),2)); c.font = BODY_FONT; c.number_format='0.00'
    c = ws3.cell(row=rr, column=10, value=round(float(row['epa_per_play']),3)); c.font = BODY_FONT; c.number_format='0.000'
ws3.freeze_panes = "A2"
for j, w in enumerate([8,7,7,13,7,15,12,15,11,10], start=1):
    ws3.column_dimensions[get_column_letter(j)].width = w

wb.save('NFL_Best_Offense_By_Year_2000-2025.xlsx')
print("saved")
