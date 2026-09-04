#!/usr/bin/env python3
"""
load_pokemon_historical_seasons.py -- imports two real completed league
seasons from the Excel workbooks this feature replaces
(fantasy-football-db/data/pokemon_historical/*.xlsx) as archived Pokemon
Draft League seasons, for local dev/testing against real historical data
rather than only hand-typed fixtures.

A THIRD workbook this league uses, Draft_League_Doc_Template_Doubles.xlsx,
is deliberately NOT imported here -- inspecting it shows it's a blank
template (its Teams sheet is unevaluated IMPORTRANGE-style array formulas
with no cached picks) rather than a played season, so there's nothing real
to import from it.

**Scope, deliberately limited** (this is dev/test seed data, not a
production requirement -- see the approved plan's "historical import
fidelity" note): imports coaches, draft picks (at their recorded point
cost, landing in pokemon_draft_pool.cost_override so historical seasons
replay at the costs that were actually paid, independent of whatever
Phase 4's usage-based formula would compute), and the regular-season
schedule with each match's winner (from the Schedule sheet's W/L columns
for the doubles workbook, or the sign of its point differential for the
singles workbook, which doesn't have explicit W/L columns). It does NOT
import per-game or per-Pokemon kill/death stats from the Stats sheets --
that would need cross-referencing three more sheets per pick and isn't
needed to exercise schedule/standings/playoffs against real data. Each
match is stored as a single confirmed 'manual'-entry game carrying just
the winner, no stats rows.

Player accounts are created per coach name (shared across both workbooks
where the same person appears in both, e.g. "Ian Heslin") with an unusable
random password -- these are dev fixtures, not real logins. The
season commissioner is set to whichever coach is listed first in that
workbook's Source sheet; the real historical commissioner isn't recorded
anywhere in the data.

Safe to re-run: skips a workbook's import entirely if a season with that
exact name already exists.

Usage:
    python3 scripts/load_pokemon_historical_seasons.py
"""
import os
import secrets
import sqlite3
import sys
from datetime import datetime

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, "data")
SQLITE_PATH = os.path.join(DATA_DIR, "app.db")
HISTORICAL_DIR = os.path.join(DATA_DIR, "pokemon_historical")

WEBAPP_DIR = os.path.join(os.path.dirname(ROOT), "webapp")
sys.path.insert(0, WEBAPP_DIR)
from app import auth  # noqa: E402 -- needs sys.path set up first
from app.pokemon_draft import pokedex  # noqa: E402


def log(msg):
    print(f"[load_pokemon_historical_seasons] {msg}")


# ---------------------------------------------------------------------
# Per-workbook layout -- these two workbooks are built from the same
# template but were filled in with different-sized rosters (6 doubles
# coaches vs 8 singles coaches), so the Teams sheet's coach-block spacing
# and pick-column offsets genuinely differ. Hardcoded here rather than
# auto-detected since this script only ever needs to read these two
# specific files, not an arbitrary future workbook.
# ---------------------------------------------------------------------

DOUBLES = {
    "file": "Scarlet_Violet_Draft_League_Doc.xlsx",
    "season_name": "Scarlet Violet Draft League (Historical Import)",
    "format_id": "historical-sv-doubles",
    "format_name": "Scarlet Violet VGC (Historical)",
    "battle_style": "doubles",
    "roster_size": 15,
    "point_budget": 70,
    "source": {"coach_col": 6, "team_col": 7, "first_row": 3},
    # (header_col, pick_row_start, pick_row_end) -- pick_name_col =
    # header_col+1, pick_cost_col = header_col+2 for this workbook.
    "team_blocks": [(3, 5, 16), (7, 5, 16), (11, 5, 16), (15, 5, 16), (19, 5, 16), (23, 5, 16)],
    "name_col_offset": 1,
    "cost_col_offset": 2,
    "schedule": {"week_col": 3, "coach1_col": 3, "wl1_col": 4, "wl2_col": 7, "coach2_col": 8,
                 "winner_from": "wl"},
}

SINGLES = {
    "file": "2024_Spring_Singles_Draft_League.xlsx",
    "season_name": "2024 Spring Singles Draft League (Historical Import)",
    "format_id": "historical-spring-singles",
    "format_name": "Smogon OU Singles (Historical)",
    "battle_style": "singles",
    "roster_size": 13,
    "point_budget": 125,
    "source": {"coach_col": 7, "team_col": 8, "first_row": 3},
    # Two row-bands (4 coaches each) sharing the same 4 column positions.
    "team_blocks": [
        (3, 5, 17), (10, 5, 17), (17, 5, 17), (24, 5, 17),
        (3, 21, 33), (10, 21, 33), (17, 21, 33), (24, 21, 33),
    ],
    "name_col_offset": 1,
    "cost_col_offset": 5,
    "schedule": {"week_col": 3, "coach1_col": 3, "diff1_col": 5, "coach2_col": 7,
                 "winner_from": "diff"},
}


def read_coaches(ws, cfg):
    """[(coach_name, team_name), ...] in Source-sheet order (row order =
    draft order in these workbooks)."""
    coaches = []
    r = cfg["first_row"]
    while True:
        coach_name = ws.cell(row=r, column=cfg["coach_col"]).value
        if not coach_name:
            break
        team_name = ws.cell(row=r, column=cfg["team_col"]).value or f"{coach_name}'s Team"
        coaches.append((coach_name.strip(), team_name.strip()))
        r += 1
    return coaches


def read_teams(ws, layout):
    """{team_name: [(pokemon_name, cost), ...]} -- keyed by team name
    (read from the block header row), which is what ties a Teams-sheet
    block back to a Source-sheet coach."""
    teams = {}
    for header_col, row_start, row_end in layout["team_blocks"]:
        team_name = ws.cell(row=row_start - 2, column=header_col).value
        if not team_name:
            continue
        team_name = team_name.strip()
        name_col = header_col + layout["name_col_offset"]
        cost_col = header_col + layout["cost_col_offset"]
        picks = []
        for r in range(row_start, row_end):
            name = ws.cell(row=r, column=name_col).value
            if not name:
                continue
            cost = ws.cell(row=r, column=cost_col).value
            picks.append((str(name).strip(), int(cost) if cost is not None else 0))
        teams[team_name] = picks
    return teams


def read_schedule(ws, cfg):
    """[(week, coach1_name, coach2_name, winner_name), ...]. winner_name
    is None for a tie (differential of exactly 0 in the singles
    workbook -- happens a couple of times in the real data)."""
    sched = cfg["schedule"]
    rows = []
    week = None
    r = 1
    max_row = ws.max_row
    while r <= max_row:
        week_cell = ws.cell(row=r, column=sched["week_col"]).value
        if isinstance(week_cell, str) and week_cell.strip().lower().startswith("week"):
            try:
                week = int(week_cell.strip().split()[-1])
            except ValueError:
                week = None
            r += 1
            continue
        coach1 = ws.cell(row=r, column=sched["coach1_col"]).value
        coach2 = ws.cell(row=r, column=sched["coach2_col"]).value
        if week is not None and coach1 and coach2 and coach1 != "Coach 1":
            if sched["winner_from"] == "wl":
                wl1 = ws.cell(row=r, column=sched["wl1_col"]).value
                winner = coach1 if wl1 == "W" else (coach2 if wl1 == "L" else None)
            else:
                diff1 = ws.cell(row=r, column=sched["diff1_col"]).value
                if diff1 is None or diff1 == 0:
                    winner = None
                else:
                    winner = coach1 if diff1 > 0 else coach2
            rows.append((week, coach1.strip(), coach2.strip(), winner.strip() if winner else None))
        r += 1
    return rows


# The workbooks hand-type regional/Tera notes in a way PokeAPI/Showdown slugs
# don't follow -- "Alolan Vulpix" needs reordering to 'vulpix-alola', and a
# trailing "(T)" is this league's own Tera-pick flag, not a form at all.
_REGIONAL_PREFIXES = {"alolan": "alola", "galarian": "galar", "hisuian": "hisui", "paldean": "paldea"}
_TAUROS_BREED_NOTES = {"fire": "blaze", "water": "aqua", "fighting": "combat"}


def normalize_pick_name(name):
    name = name.strip()
    if name.lower().endswith("(t)"):
        name = name[: -len("(t)")].strip()
    words = name.split()
    if words and words[0].lower() in _REGIONAL_PREFIXES:
        region = _REGIONAL_PREFIXES[words[0].lower()]
        rest = " ".join(words[1:])
        if "(" in rest:
            # e.g. "Paldean Tauros (Fire)" -> base "Tauros", parenthetical names
            # the breed by its in-game color, not the slug's actual breed word
            base, _, note = rest.partition("(")
            note = note.rstrip(")").strip().lower()
            breed = _TAUROS_BREED_NOTES.get(note, note)
            name = f"{base.strip()}-{region}-{breed}-breed"
        else:
            name = f"{rest}-{region}"
    # the sheet drops Ogerpon's "-mask" suffix ('ogerpon-hearthflame' rather
    # than the PokeAPI slug 'ogerpon-hearthflame-mask')
    if name.lower().startswith("ogerpon-") and not name.lower().endswith("-mask"):
        name = f"{name}-mask"
    return name


def get_or_create_user(conn, user_cache, coach_name):
    if coach_name in user_cache:
        return user_cache[coach_name]
    username = coach_name  # these workbooks' coach names are already unique
    row = conn.execute("SELECT user_id FROM users WHERE username = ?", (username,)).fetchone()
    if row is None:
        password_hash = auth.hash_password(secrets.token_urlsafe(32))
        cur = conn.execute(
            "INSERT INTO users (username, password_hash, tier) VALUES (?, ?, 'games')",
            (username, password_hash),
        )
        user_id = cur.lastrowid
    else:
        user_id = row["user_id"]
    user_cache[coach_name] = user_id
    return user_id


def import_workbook(conn, layout):
    existing = conn.execute(
        "SELECT season_id FROM pokemon_seasons WHERE name = ?", (layout["season_name"],)
    ).fetchone()
    if existing:
        log(f"'{layout['season_name']}' already imported (season_id={existing['season_id']}) -- skipping.")
        return

    path = os.path.join(HISTORICAL_DIR, layout["file"])
    if not os.path.exists(path):
        log(f"ERROR: {path} not found.")
        return
    wb = openpyxl.load_workbook(path, data_only=True)

    coaches = read_coaches(wb["Source"], layout["source"])
    teams = read_teams(wb["Teams"], layout)
    matches = read_schedule(wb["Schedule"], layout)
    log(f"{layout['file']}: {len(coaches)} coaches, "
        f"{sum(len(p) for p in teams.values())} draft picks, {len(matches)} scheduled matches parsed.")

    user_cache = {}
    conn.execute(
        """INSERT OR IGNORE INTO pokemon_formats
               (format_id, display_name, battle_style, rules_text,
                default_roster_size, default_point_budget, default_species_clause)
           VALUES (?, ?, ?, ?, ?, ?, 1)""",
        (layout["format_id"], layout["format_name"], layout["battle_style"],
         "Imported from a historical league workbook.", layout["roster_size"], layout["point_budget"]),
    )

    commissioner_id = get_or_create_user(conn, user_cache, coaches[0][0])
    cur = conn.execute(
        """INSERT INTO pokemon_seasons
               (name, format_id, commissioner_user_id, status, roster_size_cap,
                point_budget, species_clause_enabled)
           VALUES (?, ?, ?, 'draft', ?, ?, 1)""",
        (layout["season_name"], layout["format_id"], commissioner_id,
         layout["roster_size"], layout["point_budget"]),
    )
    season_id = cur.lastrowid

    coach_id_by_name = {}
    for order, (coach_name, team_name) in enumerate(coaches, start=1):
        user_id = get_or_create_user(conn, user_cache, coach_name)
        cur = conn.execute(
            "INSERT INTO pokemon_season_coaches (season_id, user_id, team_name, draft_order) "
            "VALUES (?, ?, ?, ?)",
            (season_id, user_id, team_name, order),
        )
        coach_id_by_name[coach_name] = cur.lastrowid
        # also index by team name, since read_teams() keys picks by team name
        coach_id_by_name[team_name] = cur.lastrowid

    pick_order = 0
    unmatched = set()
    for team_name, picks in teams.items():
        coach_id = coach_id_by_name.get(team_name)
        if coach_id is None:
            log(f"  no coach found for team '{team_name}' -- skipping its picks.")
            continue
        for pokemon_name, cost in picks:
            pokemon_id = pokedex.find_by_species_name(conn, normalize_pick_name(pokemon_name))
            if pokemon_id is None:
                unmatched.add(pokemon_name)
                continue
            pick_order += 1
            conn.execute(
                "INSERT OR IGNORE INTO pokemon_draft_pool (season_id, pokemon_id, cost_override) "
                "VALUES (?, ?, ?)",
                (season_id, pokemon_id, cost),
            )
            conn.execute(
                """INSERT OR IGNORE INTO pokemon_draft_picks
                       (season_id, coach_id, pokemon_id, pick_order, cost_paid)
                   VALUES (?, ?, ?, ?, ?)""",
                (season_id, coach_id, pokemon_id, pick_order, cost),
            )
            conn.execute(
                """INSERT INTO pokemon_roster_moves
                       (season_id, coach_id, pokemon_id, move_type, cost, counts_toward_fa_cap)
                   VALUES (?, ?, ?, 'draft', ?, 0)""",
                (season_id, coach_id, pokemon_id, cost),
            )
    if unmatched:
        log(f"  {len(unmatched)} pick(s) had no Pokedex match, skipped: {sorted(unmatched)[:10]}"
            f"{' ...' if len(unmatched) > 10 else ''}")

    conn.execute(
        "UPDATE pokemon_seasons SET draft_locked_at = datetime('now') WHERE season_id = ?", (season_id,)
    )
    conn.execute(
        """INSERT INTO pokemon_draft_sessions (season_id, status, started_at, completed_at)
           VALUES (?, 'complete', datetime('now'), datetime('now'))""",
        (season_id,),
    )

    match_count = 0
    for week, coach1_name, coach2_name, winner_name in matches:
        home_id = coach_id_by_name.get(coach1_name)
        away_id = coach_id_by_name.get(coach2_name)
        if home_id is None or away_id is None:
            continue
        cur = conn.execute(
            "INSERT INTO pokemon_schedule (season_id, week, coach_id_home, coach_id_away) VALUES (?, ?, ?, ?)",
            (season_id, week, home_id, away_id),
        )
        schedule_id = cur.lastrowid
        winner_coach_id = coach_id_by_name.get(winner_name) if winner_name else None
        confirmed_at = datetime.utcnow().isoformat(sep=" ", timespec="seconds") if winner_coach_id else None
        cur = conn.execute(
            """INSERT INTO pokemon_matches
                   (schedule_id, status, winner_coach_id, confirmed_at)
               VALUES (?, ?, ?, ?)""",
            (schedule_id, "confirmed" if winner_coach_id else "unreported", winner_coach_id,
             confirmed_at),
        )
        match_id = cur.lastrowid
        if winner_coach_id:
            conn.execute(
                "INSERT INTO pokemon_match_games (match_id, game_num, entry_method, winner_coach_id) "
                "VALUES (?, 1, 'manual', ?)",
                (match_id, winner_coach_id),
            )
        match_count += 1

    conn.execute("UPDATE pokemon_seasons SET status = 'archived' WHERE season_id = ?", (season_id,))
    conn.execute(
        """INSERT INTO sync_log (table_name, source, last_synced_at, row_count, notes)
           VALUES (?, ?, datetime('now'), ?, 'historical import')
           ON CONFLICT(table_name) DO UPDATE SET
               last_synced_at=datetime('now'), row_count=excluded.row_count, notes=excluded.notes""",
        (f"pokemon_season_{season_id}", layout["file"], match_count),
    )
    conn.commit()
    log(f"  imported season_id={season_id}: {len(coaches)} coaches, {pick_order} picks, "
        f"{match_count} matches confirmed.")


def main():
    if not os.path.exists(SQLITE_PATH):
        log("app.db not found -- run scripts/build_db.py first.")
        sys.exit(1)
    if not os.path.exists(os.path.join(WEBAPP_DIR, "app", "pokemon_draft")):
        log("webapp/app/pokemon_draft not found -- is this checkout missing the webapp/ directory?")
        sys.exit(1)

    conn = sqlite3.connect(SQLITE_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    for layout in (DOUBLES, SINGLES):
        import_workbook(conn, layout)
    conn.close()
    log("done.")


if __name__ == "__main__":
    main()
